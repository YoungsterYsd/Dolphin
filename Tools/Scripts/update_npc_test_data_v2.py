"""
M12 配表更新脚本 v2：让 NPC_Diapack / Condition 严格对齐 Quest_Data 各 sub_id 状态。

改动总览：
1. 触发条件表.xlsx · Condition：
   - Type 枚举新增 'Quest_StepPendingDeliver'（Param 编码 quest_id*100+sub_id）
   - 新增 cond_id=6/7/8/9 对应 q=1 sub=1/2/3/4 的 PendingDeliver

2. NPC表.xlsx · NPC_Diapack：
   - 村长（id=1）选项重写：
     R3: sub=1, "你好，村长" → 1001, Cond=0
     R4: sub=2, "我已找回证物" → 1101, Cond=6 (q=1 sub=1 PendingDeliver)
     R5: sub=3, "我已剿除史莱姆" → 1102, Cond=7 (q=1 sub=2 PendingDeliver)
     R6: sub=4, "再见" → 1500, Cond=0
     R7: sub=5, "感谢您的帮助" → 1500, Cond=3 (Quest_Finished q=1)
   - 铁匠（id=2）维持原样：sub=1, "你好，铁匠" → 1103, Cond=0

执行后会把 Excel 直接保存（in-place）。
"""
from openpyxl import load_workbook
from pathlib import Path

PROJ = Path(__file__).resolve().parents[2]
NPC_XLSX = PROJ / "Tools" / "Excel" / "NPC表.xlsx"
COND_XLSX = PROJ / "Tools" / "Excel" / "触发条件表.xlsx"

# ──────────────────────────────────────────────────────────────
# 1) 触发条件表
# ──────────────────────────────────────────────────────────────
print(f"[1/2] 更新 {COND_XLSX.name}")
wb = load_workbook(COND_XLSX)
ws = wb["Condition"]
# 行约定：R1 表头 / R2 类型 / R3+ 数据
# 当前 R3..R8（cond=1/2/3/4/5）。我们追加 cond=6/7/8/9。
# 同时把 R2 类型行的 Type 列枚举扩展为含 Quest_StepPendingDeliver
type_row = ws[2]
# Type 列（第 4 列）
type_cell = ws.cell(row=2, column=4)
print(f"  当前 Type 类型: {type_cell.value}")
type_cell.value = "Enum(Lev,Quest_Ongoing,Quest_Finished,Quest_PendingDeliver,Quest_StepPendingDeliver)"
print(f"  → 新 Type 类型: {type_cell.value}")

# 找最后一个有数据的行
max_row = ws.max_row
# 清掉可能残留的旧 cond=6+ 数据（按 id 列扫）
for r in range(3, max_row + 1):
    id_val = ws.cell(row=r, column=2).value
    if id_val is None:
        continue
    if int(id_val) >= 6:
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).value = None

# 追加 cond=6/7/8/9
new_rows = [
    (1, 6, 1, "Quest_StepPendingDeliver", 101),  # q=1 sub=1 待交付
    (1, 7, 1, "Quest_StepPendingDeliver", 102),
    (1, 8, 1, "Quest_StepPendingDeliver", 103),
    (1, 9, 1, "Quest_StepPendingDeliver", 104),
]
# 追加位置：从 R9 开始（R3-R8 已被占用）
target_row = 9
for row_data in new_rows:
    for col, val in enumerate(row_data, start=1):
        ws.cell(row=target_row, column=col).value = val
    target_row += 1

wb.save(COND_XLSX)
print(f"  ✅ Condition 已写入（追加 4 行 cond=6/7/8/9）")

# ──────────────────────────────────────────────────────────────
# 2) NPC 表（NPC_Diapack）
# ──────────────────────────────────────────────────────────────
print(f"[2/2] 更新 {NPC_XLSX.name}")
wb = load_workbook(NPC_XLSX)
ws = wb["NPC_Diapack"]

# 清掉所有数据行（保留 R1/R2）
max_row = ws.max_row
for r in range(3, max_row + 1):
    for c in range(1, ws.max_column + 1):
        ws.cell(row=r, column=c).value = None

# 写入新数据
# columns: _export, id, sub_id, Talk_Text, Dialogue_ID, Condition
new_rows = [
    # 村长
    (1, 1, 1, "你好，村长",         1001, 0),
    (1, 1, 2, "我已找回证物",       1101, 6),
    (1, 1, 3, "我已剿除史莱姆",     1102, 7),
    (1, 1, 4, "再见",               1500, 0),
    (1, 1, 5, "感谢您的帮助",       1500, 3),
    # 铁匠
    (1, 2, 1, "你好，铁匠",         1103, 0),
]

target_row = 3
for row_data in new_rows:
    for col, val in enumerate(row_data, start=1):
        ws.cell(row=target_row, column=col).value = val
    target_row += 1

wb.save(NPC_XLSX)
print(f"  ✅ NPC_Diapack 已写入（村长 5 行 + 铁匠 1 行 = 6 行）")

print("\n配表完成。下一步执行：")
print("  cd Tools/excel2Config")
print("  python main.py convert -i ../Excel/NPC表.xlsx --format csv -o ../../Data/FromExcel")
print("  python main.py convert -i ../Excel/触发条件表.xlsx --format csv -o ../../Data/FromExcel")
