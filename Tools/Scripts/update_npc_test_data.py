# -*- coding: utf-8 -*-
"""一次性测试数据更新脚本：批量更新 4 张 Excel（NPC/对话/任务/触发条件）。

不改字段结构（_export / id / sub_id / 业务列已由用户最终确认），只填真实可用范例数据。
范例剧本：村长 4 步链式任务（与 Plans/Dolphin设计/CSV配表契约_20260524.md 对齐）。

执行：python Tools/Scripts/update_npc_test_data.py
"""
from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parents[2]
EXCEL_DIR = ROOT / "Tools" / "Excel"


def _clear_data_rows(ws: Worksheet, header_rows: int = 2) -> None:
    """清空数据区（保留 R1 字段名 + R2 类型行），R3+ 全部删除。"""
    if ws.max_row > header_rows:
        ws.delete_rows(header_rows + 1, ws.max_row - header_rows)


def _write_rows(ws: Worksheet, rows: list[list]) -> None:
    """从第 3 行开始批量追加数据行。"""
    for r in rows:
        ws.append(r)


# ─────────────────────────────────────────────────────────────
# 1) NPC表.xlsx
# ─────────────────────────────────────────────────────────────

def update_npc_workbook() -> None:
    p = EXCEL_DIR / "NPC表.xlsx"
    wb = load_workbook(p)

    # NPC_Data: id, sub_id, Name, Diapack_ID, Scene, Talk_Show
    ws_data = wb["NPC_Data"]
    _clear_data_rows(ws_data)
    _write_rows(ws_data, [
        # 村长（id=1）：用对话包 1
        [1, 1, "", "村长", 1,
         "res://Content/NPC/Elder.tscn",
         "res://Content/UI/Portraits/Elder.png"],
        # 铁匠（id=2）：用对话包 2
        [1, 2, "", "铁匠", 2,
         "res://Content/NPC/Smith.tscn",
         "res://Content/UI/Portraits/Smith.png"],
    ])

    # NPC_Diapack: id, sub_id, Talk_Text, Dialogue_ID, Condition
    ws_pack = wb["NPC_Diapack"]
    _clear_data_rows(ws_pack)
    _write_rows(ws_pack, [
        # 对话包 1（村长）的 4 个选项：
        [1, 1, 1, "你好，村长",                 1001, 0],   # 任意时刻可见，闲聊
        [1, 1, 2, "我去解决了那群史莱姆",       1102, 4],   # 仅当任务1的当前 sub 待交付时显示
        [1, 1, 3, "我击败了 BOSS",              0,    3],   # 任务1全部完成后才显示（占位 Dialogue_ID=0 暂不接）
        [1, 1, 4, "再见",                        1500, 0],   # 总显示

        # 对话包 2（铁匠）的 1 个选项：
        [1, 2, 1, "你好，铁匠",                  1103, 0],
    ])

    # Condition: id, sub_id, Type, Param
    # 注：触发条件表是单独的 .xlsx；NPC表里没有 Condition sheet，跳过

    wb.save(p)
    print(f"[OK] {p.name} updated")


# ─────────────────────────────────────────────────────────────
# 2) 对话表.xlsx · 加 Branch_Cond 列 + 真实数据
# ─────────────────────────────────────────────────────────────

def update_dialogue_workbook() -> None:
    p = EXCEL_DIR / "对话表.xlsx"
    wb = load_workbook(p)
    ws = wb["Dialogue"]

    # 现状表头：_export, id, sub_id, 备注, Text, Branch_Text, Branch_ID
    # 目标加列 Branch_Cond（与 Branch_Text/Branch_ID 等长，0=无条件）
    # 检查是否已加过列，避免重复
    header_row = [c.value for c in ws[1]]
    if "Branch_Cond" not in header_row:
        ws.cell(row=1, column=8, value="Branch_Cond")
        ws.cell(row=2, column=8, value="List(Int)")
        print(f"[INFO] {p.name}: 新增 Branch_Cond 列")

    _clear_data_rows(ws)

    # 列：_export, id (graph_id), sub_id (node_id), 备注(Ignore), Text, Branch_Text, Branch_ID, Branch_Cond
    _write_rows(ws, [
        # graph 1001：村长接任务介绍（无分支，纯播放）
        [1, 1001, 1, "村长开场", "啊，年轻人，村子最近被史莱姆侵扰，你能帮忙吗？", "", "", ""],
        [1, 1001, 2, "村长收尾", "麻烦你了，先找几个证物回来给我。", "", "", ""],

        # graph 1101：第 1 步交付对话（无分支）
        [1, 1101, 1, "村长收物", "你拿来证物了？太好了，史莱姆出没在村外，去消灭它们吧。", "", "", ""],

        # graph 1102：第 2 步交付对话（含分支）
        [1, 1102, 1, "村长惊喜", "你竟然都解决了？接下来还有件事拜托你。", "", "", ""],
        [1, 1102, 2, "选项节点", "去找铁匠商量一下吧，他知道更多线索。", "{好的,等等}", "{3,4}", "{0,0}"],
        [1, 1102, 3, "玩家答应", "那就拜托你了。", "", "", ""],
        [1, 1102, 4, "玩家迟疑", "记得别拖太久。", "", "", ""],

        # graph 1103：第 3 步交付对话（铁匠）
        [1, 1103, 1, "铁匠开场", "村长让你来的？村外山洞里盘踞着一只 BOSS，去把它干掉！", "", "", ""],

        # graph 1500：村长闲聊
        [1, 1500, 1, "闲聊", "村子最近还算太平，谢谢你了。", "", "", ""],
    ])

    wb.save(p)
    print(f"[OK] {p.name} updated")


# ─────────────────────────────────────────────────────────────
# 3) 任务表.xlsx · 改字段名 + 真实数据
# ─────────────────────────────────────────────────────────────

def update_quest_workbook() -> None:
    p = EXCEL_DIR / "任务表.xlsx"
    wb = load_workbook(p)
    ws = wb["Quest_Data"]

    # 现状表头：_export, id, sub_id, 备注, Name, Desc, Kind, ID, Num, Drop_Rule_ID, Dialogue_ID
    # 目标：把 Dialogue_ID 改名 Deliver_Dialogue_ID（语义清晰化）
    header_row = [c.value for c in ws[1]]
    if "Dialogue_ID" in header_row and "Deliver_Dialogue_ID" not in header_row:
        idx = header_row.index("Dialogue_ID") + 1  # 1-based
        ws.cell(row=1, column=idx, value="Deliver_Dialogue_ID")
        print(f"[INFO] {p.name}: Dialogue_ID 列重命名为 Deliver_Dialogue_ID")

    _clear_data_rows(ws)

    # 列：_export, id, sub_id, 备注, Name, Desc, Kind(Item/Monster/NPC/Tricky), ID, Num, Drop_Rule_ID, Deliver_Dialogue_ID
    _write_rows(ws, [
        # 任务系列 1：村长 4 步串行
        [1, 1, 1, "第1步", "收集证物",   "找回 2 个史莱姆掉落的证物", "Item",    4, 2, 2001, 1101],
        [1, 1, 2, "第2步", "讨伐史莱姆", "村外消灭 3 只小怪 A",       "Monster", 1, 3, 2002, 1102],
        [1, 1, 3, "第3步", "拜访铁匠",   "去铁匠铺找他了解情况",       "NPC",     2, 1, 2003, 1103],
        [1, 1, 4, "第4步", "击败 BOSS",  "山洞里的 BOSS B 必须解决",  "Monster", 2, 1, 2004, 0],
    ])

    wb.save(p)
    print(f"[OK] {p.name} updated")


# ─────────────────────────────────────────────────────────────
# 4) 触发条件表.xlsx · 扩 Type Enum + 真实数据
# ─────────────────────────────────────────────────────────────

def update_condition_workbook() -> None:
    p = EXCEL_DIR / "触发条件表.xlsx"
    wb = load_workbook(p)
    ws = wb["Condition"]

    # 表头：_export, id, sub_id, Type, Param
    # R2 类型行需扩 Enum：Enum(Lev,Quest_Ongoing,Quest_Finished,Quest_PendingDeliver)
    cur_type = ws.cell(row=2, column=4).value
    new_type = "Enum(Lev,Quest_Ongoing,Quest_Finished,Quest_PendingDeliver)"
    if cur_type != new_type:
        ws.cell(row=2, column=4, value=new_type)
        print(f"[INFO] {p.name}: Type Enum 扩展为 {new_type}")

    _clear_data_rows(ws)

    # 列：_export, id, sub_id, Type, Param
    # 同 id 多 sub_id = AND 组合
    _write_rows(ws, [
        # Cond_ID=1：Lev >= 5
        [1, 1, 1, "Lev",                  5],

        # Cond_ID=2：任务系列 1 进行中（任意 sub_id 都算）
        [1, 2, 1, "Quest_Ongoing",        1],

        # Cond_ID=3：任务系列 1 已全部完成
        [1, 3, 1, "Quest_Finished",       1],

        # Cond_ID=4：任务系列 1 当前 sub 待交付
        [1, 4, 1, "Quest_PendingDeliver", 1],

        # Cond_ID=5：AND 组合 — Lev>=3 且 任务系列 1 全部完成
        [1, 5, 1, "Lev",                  3],
        [1, 5, 2, "Quest_Finished",       1],
    ])

    wb.save(p)
    print(f"[OK] {p.name} updated")


# ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    update_npc_workbook()
    update_dialogue_workbook()
    update_quest_workbook()
    update_condition_workbook()
    print("\n[DONE] 4 张 Excel 测试数据更新完成")
