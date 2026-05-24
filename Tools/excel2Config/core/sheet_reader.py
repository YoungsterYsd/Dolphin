"""读 .xlsx，把每张 sheet 转成结构化 RawSheet。

输出的 RawSheet 后续由 aggregator/validator 进一步处理。
本模块只做"读 Excel + 解析前两行 header"，不做单元格类型转换（那是 type_parser 的事）。
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .type_parser import (
    ParseError,
    TypeLiteralError,
    TypeSpec,
    parse_type_literal,
)


# ---------------------------------------------------------------------------
# 数据结构
# ---------------------------------------------------------------------------


# 前三列固定：A=_export, B=id, C=sub_id
FIXED_COL_NAMES: tuple[str, str, str] = ("_export", "id", "sub_id")
FIXED_COL_TYPES: tuple[str, str, str] = ("Int", "Int", "Int")
FIXED_COL_COUNT = 3


@dataclass
class FieldDef:
    """一列的定义（D 列起的用户字段）。"""

    name: str
    spec: TypeSpec
    col_idx: int  # 1-based 列号（与 Excel 一致）


@dataclass
class RawRow:
    """一行的原始单元格值（未经类型转换）。"""

    row_idx: int  # 1-based 行号
    cells: dict[str, Any]  # 字段名 → 原始单元格值


@dataclass
class HeaderIssue:
    """前两行（字段名行/类型行）的问题，作为校验错误的一种。"""

    row_idx: int  # 1 或 2
    col_idx: int
    reason: str


@dataclass
class RawSheet:
    """单张 sheet 的解析结果（数据行未做类型转换）。"""

    sheet_name: str
    source_file: str  # 仅用于错误信息
    fields: list[FieldDef] = field(default_factory=list)
    rows: list[RawRow] = field(default_factory=list)
    # 前两行检查到的问题（字段名/类型异常等），由 validator 转为 Warning/Error
    header_issues: list[HeaderIssue] = field(default_factory=list)
    # 前三列字段名/类型不匹配的 Warning（如 A1 不是 `_export`）
    fixed_col_warnings: list[HeaderIssue] = field(default_factory=list)

    @property
    def has_fatal_header_error(self) -> bool:
        """类型行解析失败属于 fatal —— 整张 sheet 都没法继续了。"""
        return any("类型行" in iss.reason for iss in self.header_issues)


# ---------------------------------------------------------------------------
# Workbook 读入口
# ---------------------------------------------------------------------------


def read_workbook(path: str | Path) -> list[RawSheet]:
    """读取整本 .xlsx，返回每张 sheet 的 RawSheet。

    - data_only=True：取公式计算后的值（避免拿到 "=A1+B1" 字符串）
    - 跳过隐藏 sheet 与 完全空白 sheet
    """
    path = Path(path)
    wb = load_workbook(filename=str(path), data_only=True, read_only=False)

    sheets: list[RawSheet] = []
    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue
        # 完全空表跳过（max_row<2 表示连标题行都没有）
        if ws.max_row is None or ws.max_row < 2:
            continue
        sheets.append(read_sheet(ws, str(path)))
    return sheets


def read_sheet(ws: Worksheet, source_file: str = "") -> RawSheet:
    """解析单张 sheet。

    步骤：
      1. 读第 1 行（字段名）+ 第 2 行（类型字面量）
      2. 校验前三列字段名/类型（不匹配仅产生 Warning，不阻断）
      3. 解析 D 列起每列的类型字面量；失败计入 header_issues
      4. 读第 3 行起所有数据行，按字段名收集单元格原值
    """
    sheet = RawSheet(sheet_name=ws.title, source_file=source_file)

    # ---- 读前两行 ----
    name_row = _read_row(ws, 1)
    type_row = _read_row(ws, 2)
    if not name_row or not type_row:
        sheet.header_issues.append(
            HeaderIssue(row_idx=1, col_idx=1, reason="缺少字段名行或类型行")
        )
        return sheet

    # ---- 校验前三列字段名 ----
    for i, expected_name in enumerate(FIXED_COL_NAMES):
        actual = _cell_str(name_row[i]) if i < len(name_row) else ""
        if actual != expected_name:
            sheet.fixed_col_warnings.append(
                HeaderIssue(
                    row_idx=1,
                    col_idx=i + 1,
                    reason=f"前三列字段名应为 `{expected_name}`，实际 `{actual}`",
                )
            )
    # ---- 校验前三列类型 ----
    for i, expected_type in enumerate(FIXED_COL_TYPES):
        actual = _cell_str(type_row[i]) if i < len(type_row) else ""
        if actual != expected_type:
            sheet.fixed_col_warnings.append(
                HeaderIssue(
                    row_idx=2,
                    col_idx=i + 1,
                    reason=f"前三列类型应为 `{expected_type}`，实际 `{actual}`",
                )
            )

    # ---- 解析 D 列起每列的类型字面量 + 字段名 ----
    user_col_count = max(len(name_row), len(type_row))
    for col_idx in range(FIXED_COL_COUNT + 1, user_col_count + 1):
        # openpyxl 行是 0-based list；前三列 [0..2]，D 列是 [3]
        name_cell = name_row[col_idx - 1] if col_idx - 1 < len(name_row) else None
        type_cell = type_row[col_idx - 1] if col_idx - 1 < len(type_row) else None

        name = _cell_str(name_cell)
        type_literal = _cell_str(type_cell)

        # 字段名 / 类型完全都空 → 跳过（用户可能留尾部空列）
        if not name and not type_literal:
            continue

        if not name:
            sheet.header_issues.append(
                HeaderIssue(
                    row_idx=1,
                    col_idx=col_idx,
                    reason=f"字段名为空但类型不为空（type={type_literal!r}）",
                )
            )
            continue

        if not type_literal:
            sheet.header_issues.append(
                HeaderIssue(
                    row_idx=2,
                    col_idx=col_idx,
                    reason=f"字段 `{name}` 类型行为空",
                )
            )
            continue

        try:
            spec = parse_type_literal(type_literal)
        except TypeLiteralError as e:
            sheet.header_issues.append(
                HeaderIssue(
                    row_idx=2,
                    col_idx=col_idx,
                    reason=f"字段 `{name}` 类型行非法：{e}",
                )
            )
            continue

        # Ignore 列：仍纳入 fields，但聚合器会跳过
        sheet.fields.append(FieldDef(name=name, spec=spec, col_idx=col_idx))

    # ---- 读数据行 ----
    if ws.max_row >= 3:
        for row_idx in range(3, ws.max_row + 1):
            raw_row = _read_row(ws, row_idx)
            if _is_empty_row(raw_row):
                continue
            cells: dict[str, Any] = {}
            # 前三列固定（不进 fields，但 cells 里也保存便于后续访问）
            cells["_export"] = raw_row[0] if len(raw_row) > 0 else None
            cells["id"] = raw_row[1] if len(raw_row) > 1 else None
            cells["sub_id"] = raw_row[2] if len(raw_row) > 2 else None
            # 用户字段
            for fdef in sheet.fields:
                idx = fdef.col_idx - 1
                cells[fdef.name] = raw_row[idx] if idx < len(raw_row) else None
            sheet.rows.append(RawRow(row_idx=row_idx, cells=cells))

    return sheet


# ---------------------------------------------------------------------------
# 工具
# ---------------------------------------------------------------------------


def _read_row(ws: Worksheet, row_idx: int) -> list[Any]:
    """读第 row_idx 行（1-based）的所有单元格值。"""
    # ws[row_idx] 返回 Cell 元组
    row = ws[row_idx]
    if not isinstance(row, tuple):
        row = (row,)
    return [c.value for c in row]


def _cell_str(v: Any) -> str:
    """把单元格值转为去前后空白的字符串；None / 空 → ''。"""
    if v is None:
        return ""
    s = str(v).strip()
    return s


def _is_empty_row(row: list[Any]) -> bool:
    """判断整行是否全空。"""
    for v in row:
        if v is None:
            continue
        if isinstance(v, str) and v.strip() == "":
            continue
        return False
    return True
